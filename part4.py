import openpyxl
from tkinter import filedialog
import tkinter as tk

def function4():
    def classify_data():
        try:
            # Load the selected Excel file
            file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
            if not file_path:
                result_text.config(state=tk.NORMAL)
                result_text.delete(1.0, tk.END)
                result_text.insert(tk.END, "File selection canceled.")
                result_text.config(state=tk.DISABLED)
                return

            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active

            # Get the user's input for the three parameters using the GUI
            user_frequency = float(frequency_entry.get())
            user_pri = float(pri_entry.get())
            user_pulse_width = float(pulse_width_entry.get())

            # Initialize variables to store the nearest matching data and row
            nearest_data = None
            nearest_row = None
            nearest_distance = float('inf')

            # Iterate through the rows in the Excel file
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) < 5:
                    # Skip rows with fewer than five columns
                    continue
                frequency, pri, pulse_width,classification,_ = row

                # Check if the classification is "known"
                if classification.lower() == "known":
                    # Calculate the Euclidean distance between user input and data in the Excel file
                    distance = ((user_frequency - frequency) ** 2 +
                                (user_pri - pri) ** 2 +
                                (user_pulse_width - pulse_width) * 2) * 0.5

                    # Check if this data is closer than the previously found nearest data
                    if distance < nearest_distance:
                        nearest_data = (frequency, pri, pulse_width)
                        nearest_row = row_num
                        nearest_distance = distance

            # Check if a nearest match was found and display the results
            result_text.config(state=tk.NORMAL)
            result_text.delete(1.0, tk.END)
            if nearest_data:
                result_text.insert(tk.END, "Nearest matching data:\n", "green")
                result_text.insert(tk.END, f"Frequency (MHz): {nearest_data[0]}\n")
                result_text.insert(tk.END, f"PRI (s): {nearest_data[1]}\n")
                result_text.insert(tk.END, f"Pulse Width (s): {nearest_data[2]}\n")
                result_text.insert(tk.END, f"Distance: {nearest_distance}\n")
                result_text.insert(tk.END, f"Row {nearest_row}: "
                                        f"{worksheet.cell(row=nearest_row, column=1).value}, "
                                        f"{worksheet.cell(row=nearest_row, column=2).value}, "
                                        f"{worksheet.cell(row=nearest_row, column=3).value}\n", "blue")
            else:
                result_text.insert(tk.END, "No matching data found.", "red")
                result_text.config(state=tk.DISABLED)
        except Exception as e:
            result_text.config(state=tk.NORMAL)
            result_text.delete(1.0, tk.END)
            result_text.insert(tk.END, f"Error: {str(e)}", "red")
            result_text.config(state=tk.DISABLED)

    # Create the main window
    window = tk.Tk()
    window.title("Data Classification")
    window.geometry("400x300")  # Set window size
    window.resizable(False, False)  # Prevent resizing

    # Create and place widgets for entering parameters
    frame = tk.Frame(window)
    frame.pack(pady=20)

    frequency_label = tk.Label(frame, text="Frequency (MHz):")
    frequency_label.grid(row=0, column=0)
    frequency_entry = tk.Entry(frame)
    frequency_entry.grid(row=0, column=1)

    pri_label = tk.Label(frame, text="PRI (s):")
    pri_label.grid(row=1, column=0)
    pri_entry = tk.Entry(frame)
    pri_entry.grid(row=1, column=1)

    pulse_width_label = tk.Label(frame, text="Pulse Width (s):")
    pulse_width_label.grid(row=2, column=0)
    pulse_width_entry = tk.Entry(frame)
    pulse_width_entry.grid(row=2, column=1)

    classify_button = tk.Button(window, text="Classify Data", command=classify_data)
    classify_button.pack()

    result_text = tk.Text(window, wrap=tk.WORD, height=10, width=40)
    result_text.pack()
    result_text.tag_config("green", foreground="green")
    result_text.tag_config("red", foreground="red")
    result_text.tag_config("blue", foreground="blue")
    result_text.config(state=tk.DISABLED)

    # Start the Tkinter main loop
    window.mainloop()