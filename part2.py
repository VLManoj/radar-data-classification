import tkinter as tk
import openpyxl
from tkinter import filedialog
from tkinter import ttk

def function2():

    def classify_data():
        try:
            # Load the Excel file
            file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
            if not file_path:
                return

            # Read the Excel file
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active

            # Define your range limits for each input
            min_frequency = float(min_frequency_entry.get())
            max_frequency = float(max_frequency_entry.get())
            min_pri = float(min_pri_entry.get())
            max_pri = float(max_pri_entry.get())
            min_pulse_width = float(min_pulse_width_entry.get())
            max_pulse_width = float(max_pulse_width_entry.get())

            # Start from the second row (skip the header)
            current_row = 2

            # Iterate through rows and classify data
            for row in worksheet.iter_rows(min_row=current_row, values_only=True):
                value1, value2, value3 = row[0], row[1], row[2]
                if (min_frequency <= value1 <= max_frequency) and (min_pri <= value2 <= max_pri) and (
                        min_pulse_width <= value3 <= max_pulse_width):
                    classification = 'known'
                else:
                    classification = 'unknown'

                # Add the classification to the Excel file
                worksheet.cell(row=current_row, column=4, value=classification)

                # Move to the next row
                current_row += 1

            # Save the modified Excel file
            workbook.save(file_path)
            result_label.config(text=f"Classification done and saved to '{file_path}'.")

        except Exception as e:
            result_label.config(text=f"Error: {str(e)} or select the file having 3 correct parameter")

    # Create the main window
    window = tk.Tk()
    window.title("Data Classification")

    window.geometry("600x400+320+200")
    window.resizable(True, True)

    lbl = tk.Label(window, text="Data Processing", font=("times new roman", 30, 'bold'), fg='red', bg='light blue')
    lbl.place(relx=0.39, rely=0.02)

    # Create a frame for the input fields and button
    frame = ttk.Frame(window, borderwidth=5, relief="solid", style="My.TFrame")
    frame.place(relx=0.1, rely=0.15, relwidth=0.8, relheight=0.7)
    style = ttk.Style()
    style.configure("My.TFrame", background="light yellow")

    # Create and place widgets for the Frequency
    ttk.Label(frame, text="Frequency (MHz):", font=("times new roman", 15)).place(relx=0.40, rely=0.05)
    min_frequency_entry = ttk.Entry(frame)
    min_frequency_entry.place(relx=0.2, rely=0.20)
    frelbl = tk.Label(frame, text="Min Range:")
    frelbl.place(relx=0.05, rely=0.20)
    max_frequency_entry = ttk.Entry(frame)
    max_frequency_entry.place(relx=0.7, rely=0.20)
    frelbl1 = tk.Label(frame, text="Max Range:")
    frelbl1.place(relx=0.55, rely=0.20)

    # Create and place widgets for the PRI
    ttk.Label(frame, text="PRI (s):", font=("times new roman", 15)).place(relx=0.42, rely=0.30)
    min_pri_entry = ttk.Entry(frame)
    min_pri_entry.place(relx=0.2, rely=0.40)
    prilbl = tk.Label(frame, text="Min Range:")
    prilbl.place(relx=0.05, rely=0.40)
    max_pri_entry = ttk.Entry(frame)
    max_pri_entry.place(relx=0.7, rely=0.40)
    prilbl1 = tk.Label(frame, text="Max Range:")
    prilbl1.place(relx=0.55, rely=0.40)

    # Create and place widgets for the pulse width
    ttk.Label(frame, text="Pulse Width (s):", font=("times new roman", 15)).place(relx=0.4, rely=0.50)
    min_pulse_width_entry = ttk.Entry(frame)
    min_pulse_width_entry.place(relx=0.2, rely=0.60)
    pwilbl = tk.Label(frame, text="Min Range:")
    pwilbl.place(relx=0.05, rely=0.60)
    max_pulse_width_entry = ttk.Entry(frame)
    max_pulse_width_entry.place(relx=0.7, rely=0.60)
    pwilbl1 = tk.Label(frame, text="Max Range:")
    pwilbl1.place(relx=0.55, rely=0.60)

    # Create a button to generate and save data
    generate_button = ttk.Button(frame, text="Classify data", command=classify_data)
    generate_button.place(relx=0.70, rely=0.75, height=40, width=130)

    # Create a style for the button hover effect
    style.configure("TButton", background="light green")
    style.map("TButton", background=[("active", "#90EE90")])

    # Create a label for displaying results
    result_label = ttk.Label(frame, text="", foreground="black", background="light yellow")
    result_label.place(relx=0.05, rely=0.85)

    # Start the Tkinter main loop
    window.mainloop()
